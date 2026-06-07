import csv
from datetime import datetime, time, timedelta
from io import BytesIO

from django.http import HttpResponse
from django.http import JsonResponse
from django.db.models import Avg, Count, Q
from django.db.models import Sum
from django.db.models.functions import ExtractHour
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook

from core.auth_utils import admin_required, staff_or_admin_required
from core.audit import log_action
from analytics.models import ForecastEvaluation
from analytics.services.forecasting import build_forecast_rows
from analytics.services.heat import compute_project_heat_scores
from analytics.services.spatial import build_spatial_heat_payload
from projects.models import Project
from records.models import PlayRecord
from records.thresholds import check_project_capacity_and_daily_threshold


def _minmax_norm(values):
    """Min-max 归一化到 [0, 1]；常量序列返回全 0。"""
    if not values:
        return []
    vmin = min(values)
    vmax = max(values)
    if vmax == vmin:
        return [0.0 for _ in values]
    return [(v - vmin) / (vmax - vmin) for v in values]


def _build_hot_score(records_qs):
    """
    计算综合热度评分（可解释）：
    - 正向：游玩记录数、重复游玩率、周转率（visits/capacity）
    - 负向：平均排队时长、停运占比（维护+关闭）
    - 归一化后加权求和，输出 0~100 分
    """
    agg = list(
        records_qs.values("project_id", "project__name", "project__capacity")
        .annotate(
            visits=Count("id"),
            avg_queue=Avg("queue_time"),
            repeat_sum=Sum("repeat_count"),
            maint_cnt=Count("id", filter=Q(status_snapshot=Project.STATUS_MAINTENANCE)),
            closed_cnt=Count("id", filter=Q(status_snapshot=Project.STATUS_CLOSED)),
        )
    )
    if not agg:
        return [], [], []

    plays = [a["visits"] or 0 for a in agg]
    repeat_rates = [
        (float(a["repeat_sum"] or 0) / float(a["visits"] or 1)) if (a["visits"] or 0) > 0 else 0.0 for a in agg
    ]
    turnovers = [
        (float(a["visits"] or 0) / float(a["project__capacity"] or 1)) if (a["project__capacity"] or 0) > 0 else 0.0
        for a in agg
    ]
    avg_queues = [float(a["avg_queue"] or 0) for a in agg]
    downtime_ratios = [
        (float((a["maint_cnt"] or 0) + (a["closed_cnt"] or 0)) / float(a["visits"] or 1)) if (a["visits"] or 0) > 0 else 0.0
        for a in agg
    ]

    n_plays = _minmax_norm(plays)
    n_repeat = _minmax_norm(repeat_rates)
    n_turn = _minmax_norm(turnovers)
    n_queue = _minmax_norm(avg_queues)
    n_down = _minmax_norm(downtime_ratios)

    # 权重可按业务口径调整。
    w_plays = 0.45
    w_repeat = 0.20
    w_turn = 0.20
    w_queue_penalty = 0.10
    w_down_penalty = 0.05

    rows = []
    for i, a in enumerate(agg):
        raw = (
            w_plays * n_plays[i]
            + w_repeat * n_repeat[i]
            + w_turn * n_turn[i]
            - w_queue_penalty * n_queue[i]
            - w_down_penalty * n_down[i]
        )
        score = round(max(0.0, min(1.0, raw)) * 100, 1)
        rows.append(
            {
                "name": a["project__name"],
                "score": score,
                "visits": int(a["visits"] or 0),
                "avg_queue": round(float(a["avg_queue"] or 0), 1),
                "repeat_rate": round(repeat_rates[i] * 100, 1),
                "turnover": round(turnovers[i] * 100, 1),
                "downtime_ratio": round(downtime_ratios[i] * 100, 1),
            }
        )

    rows.sort(key=lambda x: x["score"], reverse=True)
    labels = [r["name"] for r in rows[:8]]
    values = [r["score"] for r in rows[:8]]
    return labels, values, rows


def _pct_change(cur: float, prev: float):
    """百分比变化；当前值与前值比较，前值为 0/None 时返回 None。"""
    if prev is None or prev == 0:
        return None
    return round((cur - prev) / prev * 100, 1)


def _linear_regression_next(series):
    """一元线性回归（最小二乘）预测下一个点。"""
    n = len(series)
    if n <= 0:
        return 0.0
    if n == 1:
        return float(series[0])
    xs = list(range(n))
    ys = [float(v) for v in series]
    x_mean = sum(xs) / n
    y_mean = sum(ys) / n
    sxx = sum((x - x_mean) ** 2 for x in xs)
    if sxx == 0:
        return float(series[-1])
    sxy = sum((xs[i] - x_mean) * (ys[i] - y_mean) for i in range(n))
    b = sxy / sxx
    a = y_mean - b * x_mean
    y_next = a + b * n
    return max(0.0, float(y_next))


@staff_or_admin_required
def index(request):
    """运营看板首页，展示核心指标和图表数据。"""
    today = timezone.localdate()
    start_date_str = request.GET.get("start_date", "")
    end_date_str = request.GET.get("end_date", "")

    default_days = int(request.session.get("default_days", 7))
    default_start = today - timedelta(days=max(default_days - 1, 0))

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else default_start
    except ValueError:
        start_date = today
    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else today
    except ValueError:
        end_date = today
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    day_start = timezone.make_aware(datetime.combine(start_date, time.min))
    day_end = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), time.min))
    period_days = max((end_date - start_date).days + 1, 1)
    prev_start = day_start - timedelta(days=period_days)
    prev_end = day_start
    yoy_start = day_start - timedelta(days=365)
    yoy_end = day_end - timedelta(days=365)
    qoq_start = day_start - timedelta(days=90)
    qoq_end = day_end - timedelta(days=90)

    today_records = PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
    prev_records = PlayRecord.objects.filter(play_time__gte=prev_start, play_time__lt=prev_end)
    yoy_records = PlayRecord.objects.filter(play_time__gte=yoy_start, play_time__lt=yoy_end)
    qoq_records = PlayRecord.objects.filter(play_time__gte=qoq_start, play_time__lt=qoq_end)
    total_visits = today_records.count()
    prev_total_visits = prev_records.count()
    yoy_total_visits = yoy_records.count()
    qoq_total_visits = qoq_records.count()
    avg_queue = today_records.aggregate(v=Avg("queue_time"))["v"] or 0
    total_repeat = today_records.aggregate(v=Sum("repeat_count"))["v"] or 0
    repeat_rate = (total_repeat / total_visits) if total_visits else 0
    prev_avg_queue = prev_records.aggregate(v=Avg("queue_time"))["v"] or 0
    yoy_avg_queue = yoy_records.aggregate(v=Avg("queue_time"))["v"] or 0
    qoq_avg_queue = qoq_records.aggregate(v=Avg("queue_time"))["v"] or 0
    active_projects = Project.objects.exclude(status=Project.STATUS_CLOSED).count()
    total_projects = Project.objects.count()
    healthy_ratio = (active_projects / total_projects * 100) if total_projects else 0
    visits_change_pct = ((total_visits - prev_total_visits) / prev_total_visits * 100) if prev_total_visits else 0
    queue_change = avg_queue - prev_avg_queue
    visits_yoy_pct = _pct_change(total_visits, yoy_total_visits) if yoy_total_visits else None
    visits_qoq_pct = _pct_change(total_visits, qoq_total_visits) if qoq_total_visits else None
    avg_queue_yoy_pct = _pct_change(avg_queue, yoy_avg_queue) if yoy_avg_queue else None
    avg_queue_qoq_pct = _pct_change(avg_queue, qoq_avg_queue) if qoq_avg_queue else None

    # 周转率（turnover）：区间内“游玩记录数 / 项目承载量”做效率衡量。
    project_turnover_stats = (
        PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
        .values("project_id", "project__name", "project__capacity")
        .annotate(visits=Count("id"))
    )
    turnover_list = []
    sum_capacity = 0
    for item in project_turnover_stats:
        cap = item.get("project__capacity") or 0
        visits = item["visits"]
        sum_capacity += cap
        project_turnover = (visits / cap * 100) if cap else 0
        turnover_list.append(
            {
                "name": item["project__name"],
                "visits": visits,
                "capacity": cap,
                "turnover_rate": round(project_turnover, 1),
            }
        )

    turnover_overall_rate = (total_visits / sum_capacity * 100) if sum_capacity else 0
    turnover_rate = round(turnover_overall_rate, 1)
    turnover_top_rows = sorted(turnover_list, key=lambda x: x["turnover_rate"], reverse=True)[:5]
    turnover_labels = [x["name"] for x in turnover_top_rows]
    turnover_values = [x["turnover_rate"] for x in turnover_top_rows]

    rank_queryset = (
        PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
        .values("project__name")
        .annotate(visits=Count("id"))
        .order_by("-visits")[:8]
    )
    rank_labels = [item["project__name"] for item in rank_queryset]
    rank_values = [item["visits"] for item in rank_queryset]
    rank_rows = [
        {"name": item["project__name"], "visits": item["visits"], "avg_queue": round(item["avg_queue"] or 0, 1)}
        for item in (
            PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
            .values("project__name")
            .annotate(visits=Count("id"), avg_queue=Avg("queue_time"))
            .order_by("-visits")[:10]
        )
    ]

    analytics_score_rows = compute_project_heat_scores(start_date=start_date, end_date=end_date)
    score_rows = [
        {
            "project_id": row["project_id"],
            "name": row["project_name"],
            "score": row["score"],
            "visits": row["metrics"]["visits"],
            "avg_queue": row["metrics"]["avg_queue"],
            "repeat_rate": row["metrics"]["repeat_rate"],
            "turnover": row["metrics"]["turnover"],
            "downtime_ratio": row["metrics"].get("downtime_minutes", 0),
            "dimensions": row["dimensions"],
            "metrics": row["metrics"],
        }
        for row in analytics_score_rows
    ]
    score_labels = [row["name"] for row in score_rows[:8]]
    score_values = [row["score"] for row in score_rows[:8]]

    # 简易预测：最近 7 天，3 日均值 + 线性回归。
    prediction_rows = []
    for project in Project.objects.all():
        series = []
        for i in range(6, -1, -1):
            d = today - timedelta(days=i)
            c = PlayRecord.objects.filter(project=project, play_time__date=d).count()
            series.append(c)
        predicted_ma = round(sum(series[-3:]) / 3, 1) if series else 0
        predicted_lr = round(_linear_regression_next(series), 1) if series else 0
        predicted_best = max(predicted_ma, predicted_lr)
        capacity_risk_threshold = project.capacity * 12
        is_alert = predicted_best >= project.daily_warn_threshold or predicted_best >= capacity_risk_threshold
        prediction_rows.append(
            {
                "name": project.name,
                "predicted_next_day": predicted_ma,
                "predicted_lr": predicted_lr,
                "predicted_best": predicted_best,
                "threshold": project.daily_warn_threshold,
                "capacity_risk_threshold": capacity_risk_threshold,
                "is_alert": is_alert,
            }
        )
    prediction_rows.sort(key=lambda x: x["predicted_next_day"], reverse=True)
    alert_rows = [row for row in prediction_rows if row["is_alert"]]
    forecast_payload = build_forecast_rows(days=30, horizon=7)
    forecast_items = forecast_payload["items"][:10]
    forecast_alerts = [item for item in forecast_payload["items"] if item["alert"]]
    spatial_heat_items = build_spatial_heat_payload(days=7)["items"][:10]
    evaluation_rows = [
        {
            "project_name": item.project.name,
            "model_name": item.model_name,
            "mae": item.mae,
            "mse": item.mse,
            "r2": item.r2,
            "sample_size": item.sample_size,
        }
        for item in ForecastEvaluation.objects.select_related("project").order_by("-evaluated_at")[:10]
    ]

    hourly_counts = (
        PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
        .annotate(hour=ExtractHour("play_time"))
        .values("hour")
        .annotate(total=Count("id"))
        .order_by("hour")
    )
    hourly_map = {int(item["hour"]): item["total"] for item in hourly_counts if item["hour"] is not None}
    traffic_labels = [f"{h:02d}:00" for h in range(10, 19)]
    traffic_values = [hourly_map.get(h, 0) for h in range(10, 19)]

    status_counts = Project.objects.aggregate(
        normal=Count("id", filter=Q(status=Project.STATUS_NORMAL)),
        maintenance=Count("id", filter=Q(status=Project.STATUS_MAINTENANCE)),
        closed=Count("id", filter=Q(status=Project.STATUS_CLOSED)),
    )
    status_data = [
        {"name": "正常", "value": status_counts["normal"] or 0},
        {"name": "维护", "value": status_counts["maintenance"] or 0},
        {"name": "关闭", "value": status_counts["closed"] or 0},
    ]

    type_map = {
        Project.TYPE_THRILL: "刺激类",
        Project.TYPE_FAMILY: "亲子类",
        Project.TYPE_VIEW: "观光类",
    }
    type_ratio_data = [
        {"name": type_map.get(item["project__project_type"], item["project__project_type"]), "value": item["total"]}
        for item in (
            PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
            .values("project__project_type")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
    ]

    # 区域热力图（分区示意 3x2）。
    region_layout = {
        Project.REGION_ENTRANCE: {"x": 0, "y": 0, "name": Project.REGION_ENTRANCE},
        Project.REGION_FAMILY: {"x": 1, "y": 0, "name": Project.REGION_FAMILY},
        Project.REGION_REST: {"x": 2, "y": 0, "name": Project.REGION_REST},
        Project.REGION_THRILL: {"x": 0, "y": 1, "name": Project.REGION_THRILL},
        Project.REGION_VIEW: {"x": 1, "y": 1, "name": Project.REGION_VIEW},
        Project.REGION_CATERING: {"x": 2, "y": 1, "name": Project.REGION_CATERING},
    }
    region_code_to_label = dict(Project.REGION_CHOICES)
    region_heatmap_x_labels = ["左区", "中区", "右区"]
    region_heatmap_y_labels = ["上区", "下区"]

    region_visits = {code: 0 for code, _ in Project.REGION_CHOICES}
    project_visits = (
        PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
        .values("project_id")
        .annotate(visits=Count("id"))
    )
    project_id_list = [row["project_id"] for row in project_visits]
    projects_map = Project.objects.in_bulk(project_id_list)
    for row in project_visits:
        pid = row["project_id"]
        visits = row["visits"]
        project = projects_map.get(pid)
        if not project:
            continue
        rcode = project.effective_region()
        region_visits[rcode] = region_visits.get(rcode, 0) + visits

    region_heatmap_data = []
    region_heatmap_max = 0
    for rcode, cell in region_layout.items():
        v = region_visits.get(rcode, 0)
        region_heatmap_max = max(region_heatmap_max, v)
        region_heatmap_data.append(
            {
                "value": [cell["x"], cell["y"], v],
                "name": region_code_to_label.get(rcode, rcode),
            }
        )

    # 热度衰减周期识别：峰值后首次下降到峰值 30% 以下所需天数。
    decay_rows = []
    current_period_days = (end_date - start_date).days + 1
    decay_window_days = max(1, min(current_period_days, 30))
    decay_start_date = end_date - timedelta(days=decay_window_days - 1)
    decay_dates = [decay_start_date + timedelta(days=i) for i in range(decay_window_days)]
    # 只计算最近有数据的项目，避免大量空表。
    project_ids_in_period = (
        PlayRecord.objects.filter(play_time__date__gte=decay_start_date, play_time__date__lte=end_date)
        .values_list("project_id", flat=True)
        .distinct()
    )
    projects_for_decay = Project.objects.filter(id__in=project_ids_in_period)
    for project in projects_for_decay:
        daily_map = {
            row["play_time__date"]: row["total"]
            for row in (
                PlayRecord.objects.filter(
                    project=project,
                    play_time__date__gte=decay_start_date,
                    play_time__date__lte=end_date,
                )
                .values("play_time__date")
                .annotate(total=Count("id"))
            )
        }
        series = [daily_map.get(d, 0) for d in decay_dates]
        peak = max(series) if series else 0
        if peak <= 0:
            continue
        peak_idx = series.index(peak)
        threshold = peak * 0.3
        decay_days = None
        for j in range(peak_idx + 1, len(series)):
            if series[j] <= threshold:
                decay_days = j - peak_idx
                break
        if decay_days is None:
            decay_days = len(series) - 1 - peak_idx
        decay_rows.append(
            {
                "name": project.name,
                "peak": peak,
                "peak_day": decay_dates[peak_idx].strftime("%Y-%m-%d"),
                "decay_days": decay_days,
                "threshold": round(threshold, 1),
            }
        )
    decay_rows.sort(key=lambda x: x["peak"], reverse=True)
    decay_rows = decay_rows[:10]

    # 最近 7 天迷你趋势图数据。
    spark_days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    spark_visits_map = {
        item["play_time__date"]: item["total"]
        for item in (
            PlayRecord.objects.filter(play_time__date__gte=spark_days[0], play_time__date__lte=spark_days[-1])
            .values("play_time__date")
            .annotate(total=Count("id"))
        )
    }
    spark_queue_map = {
        item["play_time__date"]: round(item["avg"] or 0, 1)
        for item in (
            PlayRecord.objects.filter(play_time__date__gte=spark_days[0], play_time__date__lte=spark_days[-1])
            .values("play_time__date")
            .annotate(avg=Avg("queue_time"))
        )
    }
    spark_labels = [d.strftime("%m-%d") for d in spark_days]
    spark_visits = [spark_visits_map.get(d, 0) for d in spark_days]
    spark_queue = [spark_queue_map.get(d, 0) for d in spark_days]
    queue_alert_rows = []
    alerted_keys = set(request.session.get("queue_alert_logged_keys", []))
    check_time = timezone.now()
    for p in Project.objects.all().order_by("name"):
        result = check_project_capacity_and_daily_threshold(p, check_time)
        alert_key = f"{today}:{p.id}"
        if not result["allow"]:
            reasons = result["messages"]
            if alert_key not in alerted_keys:
                log_action(
                    request,
                    "queue_capacity_alert",
                    target_type="Project",
                    target_id=str(p.id),
                    message="；".join(reasons),
                )
                alerted_keys.add(alert_key)
            queue_alert_rows.append(
                {
                    "name": p.name,
                    "queue_count": p.queue_count,
                    "capacity": p.capacity,
                    "day_visits": result["day_count"],
                    "daily_warn_threshold": p.daily_warn_threshold,
                    "reasons": reasons,
                }
            )
        elif alert_key in alerted_keys:
            log_action(
                request,
                "queue_capacity_recovered",
                target_type="Project",
                target_id=str(p.id),
                message=f"{p.name} 告警恢复",
            )
            alerted_keys.discard(alert_key)

    request.session["queue_alert_logged_keys"] = sorted(alerted_keys)

    return render(
        request,
        "dashboard/index.html",
        {
            "metrics": {
                "total_visits": total_visits,
                "avg_queue": round(avg_queue, 1),
                "active_projects": active_projects,
                "total_projects": total_projects,
                "healthy_ratio": round(healthy_ratio, 1),
                "visits_change_pct": round(visits_change_pct, 1),
                "visits_yoy_pct": visits_yoy_pct,
                "visits_qoq_pct": visits_qoq_pct,
                "queue_change": round(queue_change, 1),
                "avg_queue_yoy_pct": avg_queue_yoy_pct,
                "avg_queue_qoq_pct": avg_queue_qoq_pct,
                "repeat_rate": round(repeat_rate * 100, 1),
                "turnover_rate": turnover_rate,
            },
            "rank_labels": rank_labels,
            "rank_values": rank_values,
            "traffic_labels": traffic_labels,
            "traffic_values": traffic_values,
            "status_data": status_data,
            "rank_rows": rank_rows,
            "score_labels": score_labels,
            "score_values": score_values,
            "score_rows": score_rows[:10],
            "start_date": start_date.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
            "spark_labels": spark_labels,
            "spark_visits": spark_visits,
            "spark_queue": spark_queue,
            "prediction_rows": prediction_rows[:10],
            "alert_rows": alert_rows,
            "forecast_mode": forecast_payload["mode"],
            "forecast_items": forecast_items,
            "forecast_alerts": forecast_alerts,
            "spatial_heat_items": spatial_heat_items,
            "evaluation_rows": evaluation_rows,
            "type_ratio_data": type_ratio_data,
            "turnover_labels": turnover_labels,
            "turnover_values": turnover_values,
            "top_turnover_rows": turnover_top_rows,
            "region_heatmap_x_labels": region_heatmap_x_labels,
            "region_heatmap_y_labels": region_heatmap_y_labels,
            "region_heatmap_data": region_heatmap_data,
            "region_heatmap_max": region_heatmap_max,
            "decay_rows": decay_rows,
            "queue_alert_rows": queue_alert_rows,
        },
    )


@staff_or_admin_required
def spatial_heat(request):
    days = request.GET.get("days", "7")
    try:
        days_i = max(1, min(int(days), 180))
    except ValueError:
        days_i = 7
    payload = build_spatial_heat_payload(days=days_i)
    return render(request, "dashboard/spatial_heat.html", {"payload": payload, "days": days_i})


@admin_required
def export_csv(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    query = PlayRecord.objects.select_related("project", "created_by").order_by("-play_time")
    if start_date:
        query = query.filter(play_time__date__gte=start_date)
    if end_date:
        query = query.filter(play_time__date__lte=end_date)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="play_records.csv"'
    log_action(request, "export_csv", target_type="PlayRecord", target_id="", message="导出游玩记录")
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(["项目名称", "游玩时间", "排队时长(分钟)", "重复次数", "状态快照", "备注", "录入人"])
    for item in query:
        writer.writerow(
            [
                item.project.name,
                timezone.localtime(item.play_time).strftime("%Y-%m-%d %H:%M"),
                item.queue_time,
                item.repeat_count,
                item.get_status_snapshot_display(),
                item.note,
                item.created_by.username if item.created_by else "",
            ]
        )
    return response


@admin_required
def export_xlsx(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    query = PlayRecord.objects.select_related("project", "created_by").order_by("-play_time")
    if start_date:
        query = query.filter(play_time__date__gte=start_date)
    if end_date:
        query = query.filter(play_time__date__lte=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "游玩记录"
    ws.append(["项目名称", "游玩时间", "排队时长(分钟)", "重复次数", "状态快照", "备注", "录入人"])
    for item in query:
        ws.append(
            [
                item.project.name,
                timezone.localtime(item.play_time).strftime("%Y-%m-%d %H:%M"),
                item.queue_time,
                item.repeat_count,
                item.get_status_snapshot_display(),
                item.note,
                item.created_by.username if item.created_by else "",
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="play_records.xlsx"'
    log_action(request, "export_xlsx", target_type="PlayRecord", target_id="", message="导出游玩记录")
    return response


def _parse_date_range_from_request(request, *, today):
    start_date_str = request.GET.get("start_date", "").strip()
    end_date_str = request.GET.get("end_date", "").strip()

    default_days = int(request.session.get("default_days", 7))
    default_start = today - timedelta(days=max(default_days - 1, 0))

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else default_start
    except ValueError:
        start_date = default_start

    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else today
    except ValueError:
        end_date = today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    day_start = timezone.make_aware(datetime.combine(start_date, time.min))
    day_end = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), time.min))
    return start_date, end_date, day_start, day_end


def _parse_date_range_from_request_api(request, *, today):
    """API 日期参数兼容 start/end 与 start_date/end_date。"""
    start_date_str = request.GET.get("start", "").strip() or request.GET.get("start_date", "").strip()
    end_date_str = request.GET.get("end", "").strip() or request.GET.get("end_date", "").strip()

    default_days = int(request.session.get("default_days", 7))
    default_start = today - timedelta(days=max(default_days - 1, 0))

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else default_start
    except ValueError:
        start_date = default_start

    try:
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else today
    except ValueError:
        end_date = today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    day_start = timezone.make_aware(datetime.combine(start_date, time.min))
    day_end = timezone.make_aware(datetime.combine(end_date + timedelta(days=1), time.min))
    return start_date, end_date, day_start, day_end


def _api_resp(*, code: int, message: str, data):
    return JsonResponse({"code": code, "message": message, "data": data}, json_dumps_params={"ensure_ascii": False})


@staff_or_admin_required
def api_rank(request):
    """返回热门排行数据。GET: start/end 或 start_date/end_date, limit。"""
    try:
        today = timezone.localdate()
        start_date, end_date, day_start, day_end = _parse_date_range_from_request_api(request, today=today)
        limit = int(request.GET.get("limit", 8))
        limit = max(1, min(limit, 50))

        rank_queryset = (
            PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
            .values("project__name")
            .annotate(visits=Count("id"), avg_queue=Avg("queue_time"))
            .order_by("-visits")[:limit]
        )
        labels = [item["project__name"] for item in rank_queryset]
        values = [item["visits"] for item in rank_queryset]
        rows = [
            {"name": item["project__name"], "visits": item["visits"], "avg_queue": round(item["avg_queue"] or 0, 1)}
            for item in rank_queryset
        ]
        return _api_resp(
            code=0,
            message="ok",
            data={"start_date": str(start_date), "end_date": str(end_date), "labels": labels, "values": values, "rows": rows},
        )
    except Exception as e:
        return _api_resp(code=1, message=str(e), data={})


@staff_or_admin_required
def api_hot_score(request):
    """返回综合热度评分排行。GET: start/end 或 start_date/end_date, limit。"""
    try:
        today = timezone.localdate()
        start_date, end_date, _day_start, _day_end = _parse_date_range_from_request_api(request, today=today)
        limit = int(request.GET.get("limit", 8))
        limit = max(1, min(limit, 50))

        rows = compute_project_heat_scores(start_date=start_date, end_date=end_date)
        labels = [row["project_name"] for row in rows[:limit]]
        values = [row["score"] for row in rows[:limit]]
        return _api_resp(
            code=0,
            message="ok",
            data={
                "start_date": str(start_date),
                "end_date": str(end_date),
                "labels": labels,
                "values": values,
                "rows": rows[:max(limit, 10)],
            },
        )
    except Exception as e:
        return _api_resp(code=1, message=str(e), data={})


@staff_or_admin_required
def api_traffic(request):
    """返回客流数据。GET: bucket=hour|day。"""
    try:
        today = timezone.localdate()
        start_date, end_date, day_start, day_end = _parse_date_range_from_request_api(request, today=today)
        bucket = (request.GET.get("bucket", "hour") or "hour").strip().lower()

        if bucket == "day":
            # 日期维度聚合：按天统计 play_record 数。
            day_labels = []
            day_values = []
            cur = start_date
            while cur <= end_date:
                next_day = cur + timedelta(days=1)
                v = PlayRecord.objects.filter(play_time__gte=timezone.make_aware(datetime.combine(cur, time.min))).filter(
                    play_time__lt=timezone.make_aware(datetime.combine(next_day, time.min))
                ).count()
                day_labels.append(cur.strftime("%Y-%m-%d"))
                day_values.append(v)
                cur = next_day
            return _api_resp(
                code=0,
                message="ok",
                data={"start_date": str(start_date), "end_date": str(end_date), "labels": day_labels, "values": day_values, "bucket": "day"},
            )

        # 默认 hour：10:00~18:00。
        hourly_counts = (
            PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
            .annotate(hour=ExtractHour("play_time"))
            .values("hour")
            .annotate(total=Count("id"))
            .order_by("hour")
        )
        hourly_map = {int(item["hour"]): item["total"] for item in hourly_counts if item["hour"] is not None}
        traffic_labels = [f"{h:02d}:00" for h in range(10, 19)]
        traffic_values = [hourly_map.get(h, 0) for h in range(10, 19)]
        return _api_resp(
            code=0,
            message="ok",
            data={"start_date": str(start_date), "end_date": str(end_date), "labels": traffic_labels, "values": traffic_values, "bucket": "hour"},
        )
    except Exception as e:
        return _api_resp(code=1, message=str(e), data={})


@staff_or_admin_required
def api_type_ratio(request):
    """返回项目类型占比。"""
    try:
        today = timezone.localdate()
        start_date, end_date, day_start, day_end = _parse_date_range_from_request_api(request, today=today)
        type_map = {
            Project.TYPE_THRILL: "刺激类",
            Project.TYPE_FAMILY: "亲子类",
            Project.TYPE_VIEW: "观光类",
        }

        type_ratio_data = [
            {"name": type_map.get(item["project__project_type"], item["project__project_type"]), "value": item["total"]}
            for item in (
                PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
                .values("project__project_type")
                .annotate(total=Count("id"))
                .order_by("-total")
            )
        ]
        return _api_resp(
            code=0,
            message="ok",
            data={"start_date": str(start_date), "end_date": str(end_date), "data": type_ratio_data},
        )
    except Exception as e:
        return _api_resp(code=1, message=str(e), data={})


@staff_or_admin_required
def api_predict(request):
    """返回预测与预警（基础模型，Prophet/LSTM 作为可选增强）。GET: days。"""
    try:
        history_days = int(request.GET.get("days", 30))
        history_days = max(7, min(history_days, 365))
        forecast = build_forecast_rows(days=history_days, horizon=7)
        prediction_rows = _dashboard_prediction_rows_from_forecast(forecast["items"])
        return _api_resp(
            code=0,
            message="ok",
            data={
                "history_days": history_days,
                "mode": forecast["mode"],
                "prediction_rows": prediction_rows[:10],
                "alert_rows": [item for item in prediction_rows if item["is_alert"]],
            },
        )
    except Exception as e:
        return _api_resp(code=1, message=str(e), data={})


def _dashboard_prediction_rows_from_forecast(items):
    project_map = Project.objects.in_bulk([item["project_id"] for item in items])
    rows = []
    for item in items:
        project = project_map.get(item["project_id"])
        first_day = (item.get("forecast") or [{}])[0]
        peak = item.get("peak") or first_day
        threshold = project.daily_warn_threshold if project else 0
        capacity_risk_threshold = project.capacity * 12 if project else 0
        predicted_next_day = first_day.get("predicted_visits", 0)
        predicted_peak = peak.get("predicted_visits", predicted_next_day)
        warning = item.get("warning") or peak.get("warning") or ""
        rows.append(
            {
                "project_id": item.get("project_id"),
                "name": item.get("project_name", ""),
                "predicted_next_day": predicted_next_day,
                "predicted_lr": predicted_peak,
                "predicted_best": max(predicted_next_day, predicted_peak),
                "threshold": threshold,
                "capacity_risk_threshold": capacity_risk_threshold,
                "is_alert": bool(item.get("alert")),
                "warning": warning,
                "forecast": item.get("forecast", []),
                "peak": peak,
            }
        )
    rows.sort(key=lambda row: row["predicted_best"], reverse=True)
    return rows


@staff_or_admin_required
def api_region_heatmap(request):
    """返回区域热力图数据（分区示意 3x2）。"""
    try:
        today = timezone.localdate()
        start_date, end_date, day_start, day_end = _parse_date_range_from_request_api(request, today=today)

        region_layout = {
            Project.REGION_ENTRANCE: {"x": 0, "y": 0, "name": Project.REGION_ENTRANCE},
            Project.REGION_FAMILY: {"x": 1, "y": 0, "name": Project.REGION_FAMILY},
            Project.REGION_REST: {"x": 2, "y": 0, "name": Project.REGION_REST},
            Project.REGION_THRILL: {"x": 0, "y": 1, "name": Project.REGION_THRILL},
            Project.REGION_VIEW: {"x": 1, "y": 1, "name": Project.REGION_VIEW},
            Project.REGION_CATERING: {"x": 2, "y": 1, "name": Project.REGION_CATERING},
        }
        region_code_to_label = dict(Project.REGION_CHOICES)
        region_heatmap_x_labels = ["左区", "中区", "右区"]
        region_heatmap_y_labels = ["上区", "下区"]

        region_visits = {code: 0 for code, _ in Project.REGION_CHOICES}

        project_visits = (
            PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
            .values("project_id")
            .annotate(visits=Count("id"))
        )
        project_id_list = [row["project_id"] for row in project_visits]
        projects_map = Project.objects.in_bulk(project_id_list)
        for row in project_visits:
            pid = row["project_id"]
            visits = row["visits"]
            project = projects_map.get(pid)
            if not project:
                continue
            rcode = project.effective_region()
            region_visits[rcode] = region_visits.get(rcode, 0) + visits

        region_heatmap_data = []
        region_heatmap_max = 0
        for rcode, cell in region_layout.items():
            v = region_visits.get(rcode, 0)
            region_heatmap_max = max(region_heatmap_max, v)
            region_heatmap_data.append(
                {
                    "value": [cell["x"], cell["y"], v],
                    "name": region_code_to_label.get(rcode, rcode),
                }
            )

        return _api_resp(
            code=0,
            message="ok",
            data={
                "start_date": str(start_date),
                "end_date": str(end_date),
                "region_heatmap_x_labels": region_heatmap_x_labels,
                "region_heatmap_y_labels": region_heatmap_y_labels,
                "region_heatmap_data": region_heatmap_data,
                "region_heatmap_max": region_heatmap_max,
            },
        )
    except Exception as e:
        return _api_resp(code=1, message=str(e), data={})


@staff_or_admin_required
def api_heat_decay(request):
    """返回热度衰减周期识别表格。"""
    try:
        today = timezone.localdate()
        start_date, end_date, day_start, day_end = _parse_date_range_from_request_api(request, today=today)
        current_period_days = (end_date - start_date).days + 1
        decay_window_days = max(1, min(current_period_days, 30))
        decay_start_date = end_date - timedelta(days=decay_window_days - 1)
        decay_dates = [decay_start_date + timedelta(days=i) for i in range(decay_window_days)]

        project_ids_in_period = (
            PlayRecord.objects.filter(play_time__date__gte=decay_start_date, play_time__date__lte=end_date)
            .values_list("project_id", flat=True)
            .distinct()
        )
        projects_for_decay = Project.objects.filter(id__in=project_ids_in_period)

        decay_rows = []
        for project in projects_for_decay:
            daily_map = {
                row["play_time__date"]: row["total"]
                for row in (
                    PlayRecord.objects.filter(
                        project=project,
                        play_time__date__gte=decay_start_date,
                        play_time__date__lte=end_date,
                    )
                    .values("play_time__date")
                    .annotate(total=Count("id"))
                )
            }
            series = [daily_map.get(d, 0) for d in decay_dates]
            peak = max(series) if series else 0
            if peak <= 0:
                continue
            peak_idx = series.index(peak)
            threshold = peak * 0.3
            decay_days = None
            for j in range(peak_idx + 1, len(series)):
                if series[j] <= threshold:
                    decay_days = j - peak_idx
                    break
            if decay_days is None:
                decay_days = len(series) - 1 - peak_idx

            decay_rows.append(
                {
                    "name": project.name,
                    "peak": peak,
                    "peak_day": decay_dates[peak_idx].strftime("%Y-%m-%d"),
                    "decay_days": decay_days,
                }
            )

        decay_rows.sort(key=lambda x: x["peak"], reverse=True)
        decay_rows = decay_rows[:10]
        return _api_resp(
            code=0,
            message="ok",
            data={
                "start_date": str(start_date),
                "end_date": str(end_date),
                "decay_rows": decay_rows,
            },
        )
    except Exception as e:
        return _api_resp(code=1, message=str(e), data={})


@staff_or_admin_required
def weekly_report(request):
    today = timezone.localdate()
    start_date, end_date, day_start, day_end = _parse_date_range_from_request(request, today=today)

    records_qs = PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
    total_visits = records_qs.count()
    avg_queue = records_qs.aggregate(v=Avg("queue_time"))["v"] or 0
    total_repeat = records_qs.aggregate(v=Sum("repeat_count"))["v"] or 0
    repeat_rate = (total_repeat / total_visits) if total_visits else 0

    type_map = dict(Project.TYPE_CHOICES)
    type_ratio_data = [
        {"name": type_map.get(item["project__project_type"], item["project__project_type"]), "value": item["total"]}
        for item in (
            records_qs.values("project__project_type")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
    ]

    top_projects = list(
        records_qs.values("project__name").annotate(visits=Count("id"), avg_queue=Avg("queue_time")).order_by("-visits")[:10]
    )

    project_turnover_stats = records_qs.values("project_id", "project__name", "project__capacity").annotate(visits=Count("id"))
    turnover_list = []
    sum_capacity = 0
    for item in project_turnover_stats:
        cap = item.get("project__capacity") or 0
        visits = item["visits"]
        sum_capacity += cap
        turnover_list.append({"name": item["project__name"], "turnover_rate": round((visits / cap * 100) if cap else 0, 1)})

    turnover_rate = round((total_visits / sum_capacity * 100) if sum_capacity else 0, 1)
    top_turnover_rows = sorted(turnover_list, key=lambda x: x["turnover_rate"], reverse=True)[:5]

    return render(
        request,
        "dashboard/weekly_report.html",
        {
            "start_date": start_date.strftime("%Y-%m-%d"),
            "end_date": end_date.strftime("%Y-%m-%d"),
            "metrics": {
                "total_visits": total_visits,
                "avg_queue": round(avg_queue, 1),
                "repeat_rate": round(repeat_rate * 100, 1),
                "turnover_rate": turnover_rate,
            },
            "type_ratio_data": type_ratio_data,
            "top_projects": top_projects,
            "top_turnover_rows": top_turnover_rows,
        },
    )


@admin_required
def export_weekly_xlsx(request):
    today = timezone.localdate()
    start_date, end_date, day_start, day_end = _parse_date_range_from_request(request, today=today)

    records_qs = PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
    total_visits = records_qs.count()
    avg_queue = records_qs.aggregate(v=Avg("queue_time"))["v"] or 0
    total_repeat = records_qs.aggregate(v=Sum("repeat_count"))["v"] or 0
    repeat_rate = (total_repeat / total_visits) if total_visits else 0

    type_map = dict(Project.TYPE_CHOICES)
    type_ratio_data = [
        {"name": type_map.get(item["project__project_type"], item["project__project_type"]), "value": item["total"]}
        for item in (
            records_qs.values("project__project_type")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
    ]

    top_projects = list(
        records_qs.values("project__name").annotate(visits=Count("id"), avg_queue=Avg("queue_time")).order_by("-visits")[:10]
    )

    project_turnover_stats = records_qs.values("project_id", "project__name", "project__capacity").annotate(visits=Count("id"))
    turnover_list = []
    sum_capacity = 0
    for item in project_turnover_stats:
        cap = item.get("project__capacity") or 0
        visits = item["visits"]
        sum_capacity += cap
        turnover_list.append(
            {
                "name": item["project__name"],
                "visits": visits,
                "capacity": cap,
                "turnover_rate": round((visits / cap * 100) if cap else 0, 1),
            }
        )

    turnover_rate = round((total_visits / sum_capacity * 100) if sum_capacity else 0, 1)
    top_turnover_rows = sorted(turnover_list, key=lambda x: x["turnover_rate"], reverse=True)[:5]

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "周报概览"
    ws_summary.append(["统计区间", f"{start_date} ~ {end_date}"])
    ws_summary.append(["总游玩记录数", total_visits])
    ws_summary.append(["平均排队时长(分钟)", round(avg_queue, 1)])
    ws_summary.append(["重复游玩率(%)", round(repeat_rate * 100, 1)])
    ws_summary.append(["周转率(%)", turnover_rate])

    ws_top = wb.create_sheet("Top 项目")
    ws_top.append(["排名", "项目名称", "游玩人次", "平均排队时长(分钟)"])
    for idx, p in enumerate(top_projects, start=1):
        ws_top.append([idx, p["project__name"], p["visits"], round(p["avg_queue"] or 0, 1)])

    ws_type = wb.create_sheet("项目类型占比")
    ws_type.append(["项目类型", "记录数"])
    for row in type_ratio_data:
        ws_type.append([row["name"], row["value"]])

    ws_turn = wb.create_sheet("Top 周转项目")
    ws_turn.append(["排名", "项目名称", "周转率(%)", "游玩人次", "承载量"])
    for idx, row in enumerate(top_turnover_rows, start=1):
        ws_turn.append([idx, row["name"], row["turnover_rate"], row["visits"], row["capacity"]])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="weekly_report_{start_date}_{end_date}.xlsx"'
    log_action(request, "export_weekly_xlsx", target_type="WeeklyReport", target_id="", message="导出周报 XLSX")
    return response


@admin_required
def export_weekly_html(request):
    today = timezone.localdate()
    start_date, end_date, day_start, day_end = _parse_date_range_from_request(request, today=today)

    records_qs = PlayRecord.objects.filter(play_time__gte=day_start, play_time__lt=day_end)
    total_visits = records_qs.count()
    avg_queue = records_qs.aggregate(v=Avg("queue_time"))["v"] or 0
    total_repeat = records_qs.aggregate(v=Sum("repeat_count"))["v"] or 0
    repeat_rate = (total_repeat / total_visits) if total_visits else 0

    type_map = dict(Project.TYPE_CHOICES)
    type_ratio_data = [
        {"name": type_map.get(item["project__project_type"], item["project__project_type"]), "value": item["total"]}
        for item in (
            records_qs.values("project__project_type")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
    ]

    top_projects = list(
        records_qs.values("project__name").annotate(visits=Count("id"), avg_queue=Avg("queue_time")).order_by("-visits")[:10]
    )

    project_turnover_stats = records_qs.values("project_id", "project__name", "project__capacity").annotate(visits=Count("id"))
    turnover_list = []
    sum_capacity = 0
    for item in project_turnover_stats:
        cap = item.get("project__capacity") or 0
        visits = item["visits"]
        sum_capacity += cap
        turnover_list.append({"name": item["project__name"], "turnover_rate": round((visits / cap * 100) if cap else 0, 1)})

    turnover_rate = round((total_visits / sum_capacity * 100) if sum_capacity else 0, 1)
    top_turnover_rows = sorted(turnover_list, key=lambda x: x["turnover_rate"], reverse=True)[:5]

    active_projects = Project.objects.exclude(status=Project.STATUS_CLOSED).count()
    total_projects = Project.objects.count()
    healthy_ratio = (active_projects / total_projects * 100) if total_projects else 0

    context = {
        "start_date": start_date.strftime("%Y-%m-%d"),
        "end_date": end_date.strftime("%Y-%m-%d"),
        "metrics": {
            "total_visits": total_visits,
            "avg_queue": round(avg_queue, 1),
            "repeat_rate": round(repeat_rate * 100, 1),
            "turnover_rate": turnover_rate,
        },
        "type_ratio_data": type_ratio_data,
        "top_projects": top_projects,
        "top_turnover_rows": top_turnover_rows,
        "healthy_ratio": round(healthy_ratio, 1),
    }

    html = render_to_string("dashboard/weekly_report.html", context, request=request)
    response = HttpResponse(html, content_type="text/html; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="weekly_brief_{start_date}_{end_date}.html"'
    log_action(request, "export_weekly_html", target_type="WeeklyReport", target_id="", message="导出周报 HTML")
    return response


