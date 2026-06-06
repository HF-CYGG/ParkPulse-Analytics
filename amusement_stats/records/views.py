import csv
from datetime import date as date_type, datetime as datetime_type, timedelta
from io import StringIO

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import load_workbook

from core.audit import log_action
from core.auth_utils import admin_required, staff_or_admin_required, user_is_admin
from projects.models import Project

from .forms import PlayRecordForm
from .models import PlayRecord
from .thresholds import check_project_capacity_and_daily_threshold


def _playrecord_queryset_for_user(user):
    qs = PlayRecord.objects.select_related("project", "created_by")
    if user_is_admin(user):
        return qs
    return qs.filter(created_by=user)


@staff_or_admin_required
def record_new(request):
    """游玩记录录入。"""
    record_block_alert = None
    if request.method == "POST":
        form = PlayRecordForm(request.POST)
        if form.is_valid():
            record = form.save(commit=False)
            check_result = check_project_capacity_and_daily_threshold(record.project, record.play_time)
            if not check_result["allow"]:
                record_block_alert = check_result
                for msg in check_result["messages"]:
                    messages.error(request, msg)
            else:
                record.created_by = request.user
                record.save()
                log_action(request, "record_create", target_type="PlayRecord", target_id=record.id, message=record.project.name)
                messages.success(request, "游玩记录已保存。")
                return redirect("record_new")
    else:
        form = PlayRecordForm()

    today = timezone.localdate()
    today_count = _playrecord_queryset_for_user(request.user).filter(play_time__date=today).count()

    query = _playrecord_queryset_for_user(request.user)
    filter_date = request.GET.get("filter_date", "").strip()
    filter_project = request.GET.get("filter_project", "").strip()
    filter_status = request.GET.get("filter_status", "").strip()
    if filter_date:
        query = query.filter(play_time__date=filter_date)
    if filter_project:
        query = query.filter(project_id=filter_project)
    if filter_status:
        query = query.filter(status_snapshot=filter_status)

    page_size = request.session.get("page_size", 10)
    paginator = Paginator(query, page_size)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(
        request,
        "records/record_new.html",
        {
            "form": form,
            "today_count": today_count,
            "page_obj": page_obj,
            "filter_date": filter_date,
            "filter_project": filter_project,
            "filter_status": filter_status,
            "projects": Project.objects.all(),
            "status_choices": Project.STATUS_CHOICES,
            "record_block_alert": record_block_alert,
        },
    )


@staff_or_admin_required
def record_edit(request, record_id):
    record = get_object_or_404(_playrecord_queryset_for_user(request.user), id=record_id)
    record_block_alert = None
    if request.method == "POST":
        form = PlayRecordForm(request.POST, instance=record)
        if form.is_valid():
            candidate = form.save(commit=False)
            check_result = check_project_capacity_and_daily_threshold(
                candidate.project,
                candidate.play_time,
                exclude_record_id=record.id,
            )
            if not check_result["allow"]:
                record_block_alert = check_result
                for msg in check_result["messages"]:
                    messages.error(request, msg)
            else:
                form.save()
                log_action(request, "record_update", target_type="PlayRecord", target_id=record.id, message=record.project.name)
                messages.success(request, "游玩记录已更新。")
                return redirect("record_new")
    else:
        form = PlayRecordForm(instance=record)

    return render(
        request,
        "records/record_edit.html",
        {"form": form, "record": record, "record_block_alert": record_block_alert},
    )


@staff_or_admin_required
def record_delete(request, record_id):
    record = get_object_or_404(_playrecord_queryset_for_user(request.user), id=record_id)
    if request.method == "POST":
        log_action(request, "record_delete", target_type="PlayRecord", target_id=record.id, message=record.project.name)
        record.delete()
        messages.warning(request, "游玩记录已删除。")
        return redirect("record_new")
    return render(request, "records/record_delete_confirm.html", {"record": record})


def _normalize_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    for ch in ["（", "）", "(", ")", "：", ":", "[", "]"]:
        s = s.replace(ch, "")
    s = "".join(c for c in s if c.isalnum() or "\u4e00" <= c <= "\u9fff")
    return s


def _pick_row_value(row: dict, candidates: list[str]):
    if not isinstance(row, dict):
        return None
    for cand in candidates:
        nc = _normalize_header(cand)
        for k, v in row.items():
            if _normalize_header(k) == nc:
                return v
    return None


def _parse_int(v):
    if v is None:
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        if v != v:
            return 0
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _parse_datetime(v):
    if v is None:
        return None
    if isinstance(v, datetime_type):
        if timezone.is_naive(v):
            return timezone.make_aware(v)
        return v
    if isinstance(v, (int, float)):
        if v != v:
            return None
        try:
            base = datetime_type(1899, 12, 30)
            dt = base + timedelta(days=float(v))
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt)
            return dt
        except Exception:
            return None
    if isinstance(v, date_type):
        dt = datetime_type.combine(v, datetime_type.min.time()).replace(hour=10, minute=0)
        return timezone.make_aware(dt)

    s = str(v).strip()
    if not s:
        return None

    fmts = [
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d",
        "%m/%d/%Y",
    ]
    for fmt in fmts:
        try:
            dt = datetime_type.strptime(s, fmt)
            if fmt in {"%Y-%m-%d", "%m/%d/%Y"}:
                dt = dt.replace(hour=10, minute=0, second=0)
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt)
            return dt
        except ValueError:
            continue
    return None


def _parse_status_snapshot(v):
    if v is None:
        return Project.STATUS_NORMAL
    s = str(v).strip()
    if not s:
        return Project.STATUS_NORMAL

    mapping = {
        "normal": Project.STATUS_NORMAL,
        "maintenance": Project.STATUS_MAINTENANCE,
        "closed": Project.STATUS_CLOSED,
        "正常": Project.STATUS_NORMAL,
        "维护": Project.STATUS_MAINTENANCE,
        "关闭": Project.STATUS_CLOSED,
    }
    key = s.replace(" ", "").replace("-", "").lower()
    key = key.replace("设备状态快照", "").replace("状态快照", "")
    if key in mapping:
        return mapping[key]
    if key.isdigit():
        num = int(key)
        if num == 0:
            return Project.STATUS_NORMAL
        if num == 1:
            return Project.STATUS_MAINTENANCE
        if num == 2:
            return Project.STATUS_CLOSED
    return Project.STATUS_NORMAL


def _resolve_project(project_id_raw, project_name_raw):
    project_obj = None
    project_id = None
    if project_id_raw not in (None, ""):
        try:
            project_id = int(float(project_id_raw))
        except Exception:
            project_id = None

    if project_id is not None:
        project_obj = Project.objects.filter(id=project_id).first()

    if not project_obj and project_name_raw not in (None, ""):
        project_name = str(project_name_raw).strip()
        project_obj = Project.objects.filter(name=project_name).first()
        if not project_obj:
            try:
                project_obj = Project.objects.filter(id=int(float(project_name))).first()
            except Exception:
                project_obj = None
    return project_obj


@admin_required
def record_import(request):
    """
    批量导入游玩记录：
    - 支持 CSV（UTF-8/带 BOM）
    - 支持 XLSX（第一张表，第一行为表头）
    """
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "请上传 CSV 或 XLSX 文件。")
            return redirect("record_import")

        filename = upload.name or ""
        ext = filename.split(".")[-1].lower()
        if ext not in {"csv", "xlsx"}:
            messages.error(request, "仅支持 CSV 或 XLSX。")
            return redirect("record_import")

        errors = []
        create_records: list[PlayRecord] = []
        row_no = 2

        def _parse_row_to_record(project_name, project_id, play_time_raw, queue_time_raw, repeat_count_raw, status_raw, note_raw):
            project_obj = _resolve_project(project_id, project_name)
            if not project_obj:
                raise ValueError(f"找不到项目：{project_name or project_id}")
            if project_obj.status == Project.STATUS_CLOSED:
                raise ValueError(f"项目“{project_obj.name}”已关闭，跳过")

            play_time = _parse_datetime(play_time_raw)
            if not play_time:
                raise ValueError("游玩时间解析失败")

            check_result = check_project_capacity_and_daily_threshold(
                project_obj,
                play_time,
                pending_records=create_records,
            )
            if not check_result["allow"]:
                raise ValueError("；".join(check_result["messages"]))

            queue_time = _parse_int(queue_time_raw)
            repeat_count = _parse_int(repeat_count_raw)
            status_snapshot = _parse_status_snapshot(status_raw)
            note = "" if note_raw is None else str(note_raw).strip()

            return PlayRecord(
                project=project_obj,
                play_time=play_time,
                queue_time=max(queue_time, 0),
                repeat_count=max(repeat_count, 0),
                status_snapshot=status_snapshot,
                note=note,
                created_by=request.user,
            )

        try:
            if ext == "csv":
                raw = upload.read()
                text = raw.decode("utf-8-sig", errors="ignore")
                reader = csv.DictReader(StringIO(text))
                for r in reader:
                    try:
                        record = _parse_row_to_record(
                            _pick_row_value(r, ["项目名称", "项目", "project", "project_name"]),
                            _pick_row_value(r, ["项目id", "项目ID", "project_id", "projectid"]),
                            _pick_row_value(r, ["游玩时间", "play_time", "playtime", "datetime"]),
                            _pick_row_value(r, ["排队时长(分钟)", "排队时长", "queue_time", "queue", "queue_time_min"]),
                            _pick_row_value(r, ["重复次数", "重复游玩次数", "repeat_count", "repeat", "重复游玩"]),
                            _pick_row_value(r, ["状态快照", "status_snapshot", "status"]),
                            _pick_row_value(r, ["备注", "note", "remark", "notes"]),
                        )
                        create_records.append(record)
                    except Exception as e:
                        errors.append(f"第 {row_no} 行：{e}")
                    row_no += 1
            else:
                wb = load_workbook(upload, read_only=True, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(min_row=1, values_only=True)
                headers = next(rows_iter, None)
                if not headers:
                    raise ValueError("Excel 第一行表头为空")

                for data_row in rows_iter:
                    row_dict = {}
                    for i, h in enumerate(headers):
                        row_dict[h] = data_row[i] if data_row and i < len(data_row) else None
                    if not any(v is not None and str(v).strip() for v in row_dict.values()):
                        row_no += 1
                        continue
                    try:
                        record = _parse_row_to_record(
                            _pick_row_value(row_dict, ["项目名称", "项目", "project", "project_name"]),
                            _pick_row_value(row_dict, ["项目id", "项目ID", "project_id", "projectid"]),
                            _pick_row_value(row_dict, ["游玩时间", "play_time", "playtime", "datetime"]),
                            _pick_row_value(row_dict, ["排队时长(分钟)", "排队时长", "queue_time", "queue", "queue_time_min"]),
                            _pick_row_value(row_dict, ["重复次数", "重复游玩次数", "repeat_count", "repeat", "重复游玩"]),
                            _pick_row_value(row_dict, ["状态快照", "status_snapshot", "status"]),
                            _pick_row_value(row_dict, ["备注", "note", "remark", "notes"]),
                        )
                        create_records.append(record)
                    except Exception as e:
                        errors.append(f"第 {row_no} 行：{e}")
                    row_no += 1
        except Exception as e:
            messages.error(request, f"文件解析失败：{e}")
            return render(
                request,
                "records/record_import.html",
                {"errors": errors, "success_count": 0, "fail_count": 0, "preview_count": 0},
            )

        success_count = 0
        fail_count = len(errors)
        if create_records:
            with transaction.atomic():
                PlayRecord.objects.bulk_create(create_records, batch_size=500)
                success_count = len(create_records)
                log_action(
                    request,
                    "record_import",
                    target_type="PlayRecord",
                    target_id="",
                    message=f"批量导入完成：成功 {success_count} 条，失败 {fail_count} 条",
                )

        if success_count:
            messages.success(request, f"导入成功：{success_count} 条。失败 {fail_count} 条。")
        else:
            messages.error(request, f"导入失败：没有可写入的数据（失败 {fail_count} 条）。")

        return render(
            request,
            "records/record_import.html",
            {
                "errors": errors,
                "success_count": success_count,
                "fail_count": fail_count,
                "preview_count": success_count,
            },
        )

    return render(request, "records/record_import.html", {"errors": [], "success_count": 0, "fail_count": 0, "preview_count": 0})


@admin_required
def record_import_template_csv(request):
    """提供 CSV 导入模板下载。"""
    if request.method != "GET":
        return redirect("record_import")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="record_import_template.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(["项目名称", "游玩时间", "排队时长(分钟)", "重复次数", "状态快照", "备注", "录入人"])
    writer.writerow(["过山车", "2026-04-15 10:00", 15, 1, "正常", "示例数据", "（可留空）"])
    return response
